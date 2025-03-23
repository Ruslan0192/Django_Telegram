import datetime
import os
from pathlib import Path

from aiogram import Bot, types
from aiogram.types import FSInputFile, InputMediaPhoto, LabeledPrice
from aiogram.fsm.context import FSMContext

import openpyxl

from config import settings, logger

from database.orm_query import *

from keyboards.inline import *

from utils.paginator import Paginator, def_pages

p = Path(os.getcwd())
IMAGE_PATH = p.parents[0]


async def def_clear_message(bot: Bot, state: FSMContext, telegram_id: int):
    # удаляю все сообщения кроме основного
    data_state = await state.get_data()
    messages_show = data_state['messages_show']

    if messages_show:
        messages_delete_values = list(messages_show.values())
        for messages_delete in messages_delete_values:
            await bot.delete_message(telegram_id, messages_delete[1])
        await state.update_data({'messages_show': None})

    await state.update_data({'get_message': None})


# ***********************************************************************************************
async def get_menu_content(message: types.Message,
                           session: AsyncSession,
                           bot: Bot,
                           state: FSMContext,
                           telegram_id: int,
                           level: int,
                           data_int1: int,
                           data_int2: int,
                           data_int3: int,
                           data_int4: int,
                           data_string: str,
                           page: int):

    if level == 0:
        return await def_choice_catalog(session, state, bot, level, telegram_id)
    if level == 1:
        return await def_choice_category(session, state, level, data_int1, page)
    if level == 2:
        return await def_choice_under_category(bot, session, state, telegram_id, level, data_int1, page)
    if level == 3:
        return await def_choice_product(message, session, state, telegram_id, level, data_int1)
    if level == 4:
        return await def_choice_product_qty(bot, session, state, telegram_id, level,
                                            data_int1, data_int2, data_int3, data_int4)

    if level == 5:
        return await def_show_order(message, bot, session, state, telegram_id, level)
    if level == 6:
        return await def_show_order_qty(bot, session, state, telegram_id, level,
                                        data_int1, data_int2, data_int3, data_int4)
    if level == 7:
        return await def_address_order(bot, session, state, telegram_id, level, data_int1)
    if level == 8:
        return await def_payment(bot, session, state, telegram_id, level, data_int1, data_string)

    if level == 10:
        return await def_orders_paid(bot, session, state, telegram_id, level)
    if level == 11:
        return await def_show_order_paid(message, bot, session, state, telegram_id, level, data_int1)

    if level == 20:
        return await def_choice_faq(bot, state, telegram_id, level)
    if level == 21:
        return await def_my_faq(session, telegram_id, level)
    if level == 22:
        return await def_show_faq(bot, session, state, telegram_id, level)
    if level == 23:
        return await def_new_question_faq(session, state, level, data_int1)
    return None, None


# ********************************************************************************************
# level == 0
async def def_choice_catalog(session: AsyncSession, state: FSMContext, bot: Bot,
                             level: int, telegram_id: int):
    # сброс и вывод каталога
    await def_clear_message(bot, state, telegram_id)

    catalogs = await orm_get_catalogs(session)

    text = 'Выберите каталог:'
    reply_markup = get_choice_catalog_btns(level=level, data_list=catalogs)
    return text, reply_markup


# level == 1
async def def_choice_category(session: AsyncSession, state: FSMContext,
                              level: int, catalog_id: int, page: int):
    # вывод категорий
    # для возврата
    await state.update_data({'catalog_id': catalog_id})

    categories = await orm_get_categories(session, catalog_id)
    if not categories:
        text = 'Категории отстутствуют.'
        reply_markup = get_empty_data_btns(level=level)
        return text, reply_markup

    paginator = Paginator(categories, page)
    pagination_btns = def_pages(paginator)
    category = paginator.get_page()

    text = 'Выберите категорию:'
    reply_markup = get_choice_category_btns(level=level,
                                            page=page,
                                            id_value=catalog_id,
                                            pagination_btns=pagination_btns,
                                            data_list=category)
    return text, reply_markup


# level == 2
async def def_choice_under_category(bot: Bot, session: AsyncSession, state: FSMContext,
                                    telegram_id: int, level: int, categories_id: int, page: int):
    # вывод подкатегорий
    data_state = await state.get_data()
    # для возврата
    catalog_id = data_state['catalog_id']
    await def_clear_message(bot, state, telegram_id)

    under_categories = await orm_get_under_categories(session, categories_id)
    if not under_categories:
        text = 'Подкатегории отстутствуют.'
        reply_markup = get_empty_data_btns(level=level, back_id=catalog_id)     # для возврата
        return text, reply_markup

    # для возврата
    await state.update_data({'categories_id': categories_id})

    paginator = Paginator(under_categories, page)
    pagination_btns = def_pages(paginator)
    under_category = paginator.get_page()

    text = 'Выберите подкатегорию:'
    reply_markup = get_choice_category_btns(level=level,
                                            page=page,
                                            id_value=categories_id,
                                            back_id=catalog_id,      # для возврата
                                            pagination_btns=pagination_btns,
                                            data_list=under_category)
    return text, reply_markup


# level == 3
async def def_choice_product(message: types.Message, session: AsyncSession, state: FSMContext,
                             telegram_id: int, level: int, undercategories_id: int):
    # вывод товаров
    data_state = await state.get_data()
    # для возврата
    categories_id = data_state['categories_id']

    products = await orm_get_products(session, undercategories_id)
    if not products:
        text = 'В этой подкатегории нет товаров.'
        reply_markup = get_empty_data_btns(level=level, back_id=categories_id)  # для возврата
        return text, reply_markup

    await state.update_data({'undercategories_id': undercategories_id})     # для возврата

    # определяю id заказа, все корзины добавляю на него
    order_id, carts = await orm_get_order_current(session, telegram_id)
    if carts:
        carts_product = {}
        for cart in carts:
            carts_product.update({cart.product_id: [cart.id, cart.quantity, cart.ammont]})
    else:
        carts_product = None

    messages_show = {}
    for product in products:
        # вывожу по одной карточке
        # проверяю корзину
        if carts_product and product.id in carts_product.keys():
            # в корзине есть товар и он в данной подкатегории
            cart = carts_product[product.id]
            cart_id = cart[0]
            quantity = cart[1]
            ammont = cart[2]
            flags_cart = True
        else:
            cart_id = 0
            quantity = 0
            ammont = 0
            flags_cart = False

        await def_new_card_product(message, product, messages_show, level+1, order_id,
                                   cart_id, quantity, ammont, flags_cart)

    # вывожу последнее сообщение навигации
    reply_markup = get_choice_product_last_btns(level=level+1, back_id=categories_id)
    message_show = await message.answer(text='Выберите действие:', reply_markup=reply_markup, parse_mode='Markdown')
    messages_show.update({'control_message': ['text', message_show.message_id]})

    await state.update_data({'messages_show': messages_show})
    return 'Выберите товар:', None


async def def_new_card_product(message: types.Message, product, messages_show: dict, level: int, order_id: int,
                               cart_id: int, quantity: int, ammont: int, flags_cart: bool, flags_views: bool = False):
    # запуск карточки продукта
    text, reply_markup = def_product_card(product.id, product.name, product.description, product.price,
                                          level, order_id, cart_id, quantity, ammont, flags_cart, flags_views)
    # проверка на наличие фото
    if product.image:
        photo_file = FSInputFile(os.path.join(IMAGE_PATH, product.image))
        message_show = await message.answer_photo(photo=photo_file,
                                                  caption=text,
                                                  reply_markup=reply_markup,
                                                  parse_mode='Markdown'
                                                  )
        messages_show.update({product.id: ['photo', message_show.message_id]})
    else:
        message_show = await message.answer(text=text, reply_markup=reply_markup, parse_mode='Markdown')
        messages_show.update({product.id: ['text', message_show.message_id]})


def def_product_card(product_id: int, name: str, description: str, price: float, level: int, order_id: int,
                     cart_id: int, quantity: int, ammont: float, flags_cart: bool, flags_views: bool = False):
    # вывод карточки товара
    text = f'*{name}*\n' \
           f'*Описание*: _{description}_\n' \
           f'*Цена:* _{price}_\n'
    if flags_cart:
        text += f'\n*В корзине: {quantity} шт, на сумму: {ammont} руб.*'

    if flags_views:
        text += f'\n*В заказе: {quantity} шт, на сумму: {ammont} руб.*'

    reply_markup = get_choice_product_btns(level=level,
                                           product_id=product_id,
                                           order_id=order_id,
                                           cart_id=cart_id,
                                           flags_cart=flags_cart,
                                           quantity=quantity,
                                           flags_views=flags_views)
    return text, reply_markup


# level == 4
async def def_choice_product_qty(bot: Bot, session: AsyncSession, state: FSMContext, telegram_id: int,
                                 level: int, product_id: int, order_id: int, cart_id: int, quantity: int):
    product = await orm_get_product(session, product_id)
    if quantity and cart_id:
        # кол-во больше 0, корзина есть, просто меняю кол-во
        ammont = product.price * quantity
        await orm_change_cart_qty(session, cart_id, quantity, ammont)
        flags_cart = True
    elif quantity:
        # есть кол-во, значит нет корзины. Создаю запись в БД
        ammont = product.price * quantity
        cart_id = await orm_add_cart(session, order_id, product_id, quantity, ammont)
        flags_cart = True
    else:
        # корзина была, но кол-во упало до 0, удаляю
        await orm_delete_cart(session, cart_id)
        ammont = 0
        flags_cart = False

    # обновляю карточку
    return await def_change_card_product(bot, state, telegram_id, product,
                                         level, order_id, cart_id, quantity, ammont, flags_cart)


async def def_change_card_product(bot: Bot, state: FSMContext, telegram_id: int, product, level: int,
                                  order_id: int, cart_id: int, quantity: int, ammont: float, flags_cart: bool):
    # обновляю карточку
    text, reply_markup = def_product_card(product.id, product.name, product.description, product.price,
                                          level, order_id, cart_id, quantity, ammont, flags_cart)

    data_state = await state.get_data()
    messages_show = data_state['messages_show']
    message_show = messages_show[product.id][1]

    if product.image:
        photo_file = FSInputFile(os.path.join(IMAGE_PATH, product.image))
        media = InputMediaPhoto(media=photo_file, caption=text, parse_mode='Markdown')
        await bot.edit_message_media(media=media,
                                     chat_id=telegram_id,
                                     message_id=message_show,
                                     reply_markup=reply_markup)
    else:
        await bot.edit_message_text(text=text,
                                    chat_id=telegram_id,
                                    message_id=message_show,
                                    reply_markup=reply_markup,
                                    parse_mode='Markdown')
    return None, None


# level == 5
async def def_show_order(message: types.Message, bot: Bot, session: AsyncSession, state: FSMContext,
                         telegram_id: int, level: int):
    # Заказ
    await def_clear_message(bot, state, telegram_id)

    # определяю id заказа, все корзины добавляю на него
    order_id, carts = await orm_get_order_current(session, telegram_id)

    if not carts:
        return 'Корзина пуста!', get_catalog_btns()

    messages_show = {}
    for cart in carts:
        product = await orm_get_product(session, cart.product_id)
        await def_new_card_product(message, product, messages_show, level+1, order_id,
                                   cart.id, cart.quantity, cart.ammont, True)

    # вывожу последнее сообщение навигации
    reply_markup = get_show_order_last_btns(level=level+1, order_id=order_id)
    message_show = await message.answer(text='Выберите действие:', reply_markup=reply_markup, parse_mode='Markdown')
    messages_show.update({'control_message': ['text', message_show.message_id]})

    await state.update_data({'messages_show': messages_show})
    return 'Корзина:', None


# level == 6
async def def_show_order_qty(bot: Bot, session: AsyncSession, state: FSMContext, telegram_id: int,
                             level: int, product_id: int, order_id: int, cart_id: int, quantity: int):
    product = await orm_get_product(session, product_id)
    if quantity:
        # кол-во больше 0, корзина есть, просто меняю кол-во
        ammont = product.price * quantity
        await orm_change_cart_qty(session, cart_id, quantity, ammont)
        flags_cart = True

        # обновляю карточку
        return await def_change_card_product(bot, state, telegram_id, product,
                                             level, order_id, cart_id, quantity, ammont, flags_cart)

    # корзина была, но кол-во упало до 0, удаляю
    await orm_delete_cart(session, cart_id)

    # проверяю есть что в заказе
    # определяю id заказа, все корзины добавляю на него
    order_id, carts = await orm_get_order_current(session, telegram_id)

    if not carts:
        await def_clear_message(bot, state, telegram_id)
        return 'Корзина пуста!', get_catalog_btns()

    # удаляю карточку из заказа, с подчисткой в каталоге карточек
    data_state = await state.get_data()
    messages_show = data_state['messages_show']
    delete_message = messages_show.pop(product.id)
    await bot.delete_message(telegram_id, delete_message[1])
    await state.update_data({'messages_show': messages_show})
    return None, None


# level == 7
async def def_address_order(bot: Bot, session: AsyncSession, state: FSMContext,
                            telegram_id: int, level: int, order_id: int):
    # Заказ, ожидание адреса
    await def_clear_message(bot, state, telegram_id)
    await state.update_data({'get_message': 'new_address'})
    await state.update_data({'order_id': order_id})

    orders = await orm_get_orders_for_address(session, telegram_id)
    if orders:
        if len(orders) == 1:
            if orders[0].address:
                # выведу все предыдущие адреса доставки, учитывая новый заказ (если адрес был уже введен)
                reply_markup = get_choice_address_btns(level=level, orders=orders, order_id=order_id)
                return 'Выберите адрес из списка или введите новый:', reply_markup
        else:
            # выведу все предыдущие адреса доставки, учитывая новый заказ (если адрес был уже введен)
            reply_markup = get_choice_address_btns(level=level, orders=orders, order_id=order_id)
            return 'Выберите адрес из списка или введите новый:', reply_markup
    return 'Введите адрес доставки:', get_empty_data_btns(level=level - 1)


# level == 8
async def def_payment(bot: Bot, session: AsyncSession, state: FSMContext, telegram_id: int,
                      level: int,  order_id: int, address: str):
    # оплата
    await state.update_data({'get_message': None})

    order, carts = await orm_get_order_id(session, order_id)

    ammont = 0      # итого
    for cart in carts:
        ammont += cart.ammont
    # записываю итого и адрес доставки
    await orm_change_order_amm_adr(session, order_id, ammont, address)

    current_datetime = datetime.datetime.now()
    str_current_datetime = current_datetime.strftime("_%d-%m-%Y_%H-%M")
    send_payload = str(telegram_id) + str_current_datetime

    amount = int(ammont * 100)
    prices = [LabeledPrice(label="оплата за товар", amount=amount)]

    messages_show = {}
    message_show = await bot.send_invoice(chat_id=telegram_id,
                                          title="счет",
                                          description="оплата в онлайн магазине",
                                          payload=send_payload,
                                          provider_token=settings.SHOP_API_TOKEN,
                                          currency='RUB',
                                          prices=prices,
                                          need_phone_number=True,
                                          )

    messages_show.update({'control_message': ['text', message_show.message_id]})
    await state.update_data({'messages_show': messages_show})
    await state.update_data({'amount': amount})

    # временно, для тестовой оплаты. Заменить на 'Оплатите счет:'
    text = "Внимание! \nОплата производится в тестовом режиме! \n" \
           "Для оплаты используйте данные тестовой карты:\n1111 1111 1111 1026\n12/26 CVC 000.\n\n"\
           "В тест режиме оплата в пределах 100 - 1000  рублей."
    return text, get_user_payment_btns(level=level)


# level == 10
async def def_orders_paid(bot: Bot, session: AsyncSession, state: FSMContext, telegram_id: int, level: int):
    await def_clear_message(bot, state, telegram_id)

    # оплаченные заказы
    orders = await orm_get_orders_paid(session, telegram_id)
    if orders:
        reply_markup = get_choice_orders_btns(level=level, orders=orders)
        return 'Выберите заказ:', reply_markup
    return 'Оплаченные заказы отсутствуют.', get_catalog_btns()


# level == 11
async def def_show_order_paid(message: types.Message, bot: Bot, session: AsyncSession, state: FSMContext,
                              telegram_id: int, level: int, order_id: int):
    # Заказ оплаченный
    await def_clear_message(bot, state, telegram_id)

    # определяю id заказа, все корзины добавляю на него
    order, carts = await orm_get_order_id(session, order_id)

    messages_show = {}
    for cart in carts:
        product = await orm_get_product(session, cart.product_id)
        await def_new_card_product(message, product, messages_show, level, order_id,
                                   cart.id, cart.quantity, cart.ammont, False, True)

    # вывожу последнее сообщение навигации
    reply_markup = get_show_order_paid_last_btns(level=level)
    message_show = await message.answer(text='Выберите действие:', reply_markup=reply_markup, parse_mode='Markdown')
    messages_show.update({'control_message': ['text', message_show.message_id]})
    await state.update_data({'messages_show': messages_show})

    return f'Заказ от {order.date_paid}:', None


async def job_with_excel(session: AsyncSession, telegram_id: int, order_id: int):
    # работа с выгрузкой заказа в таблицу
    if not os.path.exists('excel'):
        os.mkdir('excel')
    path_file = 'excel/report.xlsx'
    if not os.path.exists(path_file):
        # файла нет, создаю
        wb = openpyxl.Workbook()
        sheet = wb.active
        columns = ('Имя', 'ID telegram', 'Товар', 'Количество', 'Цена', 'Итого', 'Номер ПП', 'Дата платежа')
        sheet.append(columns)
        wb.save(filename=path_file)
        logger.info("Бот сформировал excel!", telegram_id=0)

    # подготавливаю данные к записи
    user = await orm_get_user_telegram(session, telegram_id)
    order, carts = await orm_get_order_id(session, order_id)

    # дописываю в файл
    wb = openpyxl.load_workbook(path_file)
    sheet = wb.active

    if len(carts) == 1:
        product = await orm_get_product(session, carts[0].product_id)
        columns = [user.name,
                   user.telegram_id,
                   product.name,
                   carts[0].quantity,
                   product.price,
                   order.ammont,
                   order.payment_number,
                   order.date_paid]
        sheet.append(columns)

    else:
        row_start = sheet.max_row + 1           # свободная строка
        row_end = row_start + len(carts) - 1    # последняя строка для объединения ячеек

        sheet.merge_cells(f'A{row_start}:A{row_end}')
        sheet[f'A{row_start}'] = user.name

        sheet.merge_cells(f'B{row_start}:B{row_end}')
        sheet[f'B{row_start}'] = user.telegram_id

        i = row_start
        for cart in carts:
            product = await orm_get_product(session, cart.product_id)
            sheet[f'C{i}'] = product.name
            sheet[f'D{i}'] = cart.quantity
            sheet[f'E{i}'] = product.price
            i += 1

        sheet.merge_cells(f'F{row_start}:F{row_end}')
        sheet[f'F{row_start}'] = order.ammont

        sheet.merge_cells(f'G{row_start}:G{row_end}')
        sheet[f'G{row_start}'] = order.payment_number

        sheet.merge_cells(f'H{row_start}:H{row_end}')
        sheet[f'H{row_start}'] = order.date_paid

    wb.save(path_file)
    logger.info("Заказ добавлен в excel!", telegram_id=telegram_id)


# level == 20
async def def_choice_faq(bot: Bot, state: FSMContext, telegram_id: int, level: int):
    await def_clear_message(bot, state, telegram_id)
    return f'Выберите действие.', get_choice_faq_btns(level=level)


# level == 21
async def def_my_faq(session: AsyncSession, telegram_id: int, level: int):
    questions = await orm_get_questions_user(session, telegram_id)
    if not questions:
        return 'У вас не было вопросов.', get_empty_data_btns(level=level)

    text = '*Ваша переписка:*\n'
    for question in questions:
        if question.answer:
            answer = question.answer
        else:
            answer = ''
        text += f'\n*Вопрос:* _{question.question}_\n' \
                f'*Ответ:* _{answer}_\n'

    return text, get_empty_data_btns(level=level)


# level == 22
async def def_show_faq(bot: Bot, session: AsyncSession, state: FSMContext, telegram_id: int, level: int):
    await def_clear_message(bot, state, telegram_id)
    await state.update_data({'get_message': 'new_question'})

    questions = await orm_get_questions(session)
    if not questions:
        return 'Напишите свой вопрос.', get_catalog_btns()

    reply_markup = get_choice_question_btns(level=level, questions=questions)

    return f'Выбирете интересующий вопрос или напишите свой.', reply_markup


# level == 23
async def def_new_question_faq(session: AsyncSession, state: FSMContext, level: int, question_id: int):
    await state.update_data({'get_message': None})

    question = await orm_get_question(session, question_id)

    if question.answer:
        answer = question.answer
    else:
        answer = ''
    text = f'\n*Ваш вопрос:* _{question.question}_\n' \
           f'*Ответ:* _{answer}_\n'

    return text, get_empty_data_btns(level=level)
